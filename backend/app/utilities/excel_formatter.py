from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd

# Step 1: Format the Excel sheet
def format_excel_sheet(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Set column widths and text wrap
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)  # Enable text wrap
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted workbook
    formatted_excel_path = excel_path.replace(".xlsx", "_formatted.xlsx")
    wb.save(formatted_excel_path)
    return formatted_excel_path

# Step 2: Convert Excel to PDF (conceptual, requires intermediate HTML step)
# def convert_excel_to_pdf(excel_path, pdf_path):
#     # Conversion from Excel to HTML (not directly supported here, but you can use other tools/libraries)
#     # Then, convert HTML to PDF using WeasyPrint or similar
#     pass  # Placeholder for conversion logic

# Assuming 'lesson_planner.py' and 'complete_lesson_plan.xlsx' are prepared
excel_path = "complete_lesson_plan.xlsx"
formatted_excel_path = format_excel_sheet(excel_path)

# # PDF path
# pdf_path = "path/to/complete_lesson_plan.pdf"
# convert_excel_to_pdf(formatted_excel_path, pdf_path)
